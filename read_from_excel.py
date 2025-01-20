import openpyxl
import random
from song import Song

def get_row_data(row, dataframe1):
    row_data = []
    for col in dataframe1.iter_cols(1, dataframe1.max_column - 2):
        row_data.append(col[row].value)
    return row_data

def read_excel(filename):
    songs = []
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active

    for row in range(1, dataframe1.max_row):
        parsed_song = get_row_data(row, dataframe1)
        song = Song(row, parsed_song[0], parsed_song[1], parsed_song[2], parsed_song[3])
        if not song.is_empty():
            song.generate_qr()
            songs.append(song)
    return sorted(songs, key=lambda x: random.random())